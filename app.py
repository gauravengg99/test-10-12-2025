# app.py (Python 3.14)
from __future__ import annotations
import os
import csv
import mimetypes
import pathlib
import logging
from datetime import datetime
from typing import Optional
from flask import Flask, request, jsonify, send_file
from dotenv import load_dotenv

# Excel writer
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load environment variables from .env (if present)
load_dotenv()

# ---------- Configuration (via env) ----------
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "5000"))
DEBUG = os.getenv("DEBUG", "false").lower() in ("1", "true", "yes")

# Where PDFs are stored (relative to this file)
BASE_DIR = pathlib.Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"

# Submissions log CSV (created/append)
SUBMISSIONS_CSV = BASE_DIR / "submissions.csv"
# Submissions Excel file
SUBMISSIONS_XLSX = BASE_DIR / "submissions.xlsx"

# Friendly key -> filename mapping (exact filename inside assets/)
PDF_KEY_MAP = {
    "air-cool": "Air-Cool .pdf",
    "cutter-compactor": "Cutter compactor.pdf",
    "force-feeder": "Force Feeder GE-RE-V Series .pdf",
    "dry-wash": "DRY WASH .pdf",
    "size-reduction": "Size reduction Equipments.pdf",
    "palletizing": "Palletizing Equipement  .pdf",
    # add more mappings as needed
}

# Allowed PDF extensions
ALLOWED_PDF_EXTS = {".pdf", ".PDF"}

# Small safety limit for mobile string length
MAX_FIELD_LEN = 300

# Configure logging
logging.basicConfig(level=logging.DEBUG if DEBUG else logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")

app = Flask(__name__, static_folder=".", static_url_path="/")


# ---------- Utility functions ----------
def safe_str(s: Optional[str]) -> str:
    return (s or "").strip()


def validate_submission(name: str, email: str, mobile: str) -> Optional[str]:
    if not name:
        return "Name is required."
    if not email:
        return "Email is required."
    if not mobile:
        return "Mobile number is required."
    # basic length checks
    if len(name) > MAX_FIELD_LEN or len(email) > MAX_FIELD_LEN or len(mobile) > MAX_FIELD_LEN:
        return "One or more fields too long."
    return None


def find_pdf_by_key(key: Optional[str] = None) -> Optional[pathlib.Path]:
    """
    Find a PDF file inside ASSETS_DIR.
    - First check PDF_KEY_MAP for exact filename.
    - Then fall back to substring-match heuristics.
    - Finally return the first PDF found.
    """
    if not ASSETS_DIR.exists():
        logging.error("Assets directory %s does not exist", ASSETS_DIR)
        return None

    wanted = (key or "").strip().lower()

    # list of pdf files
    pdf_files = [p for p in ASSETS_DIR.iterdir() if p.is_file() and p.suffix.lower() in ALLOWED_PDF_EXTS]
    if not pdf_files:
        logging.warning("No PDF files found in assets directory.")
        return None

    # 1) check mapping
    if wanted:
        mapped = PDF_KEY_MAP.get(wanted)
        if mapped:
            candidate = ASSETS_DIR / mapped
            if candidate.exists() and candidate.suffix.lower() in ALLOWED_PDF_EXTS:
                logging.info("Mapped key '%s' -> file '%s'", wanted, mapped)
                return candidate
            else:
                logging.warning("Mapped file for key '%s' -> '%s' not found.", wanted, mapped)

    # 2) substring match (case-insensitive) against filenames
    if wanted:
        for p in pdf_files:
            if wanted in p.name.lower():
                logging.info("Substring match key '%s' -> file '%s'", wanted, p.name)
                return p

    # 3) fallback heuristics (common variants)
    fallback_terms = ["air", "air-cool", "air cool", "air-coolseries", "aircool",
                      "cutter", "compactor", "force", "feeder", "dry", "wash", "size", "pallet"]
    for term in fallback_terms:
        for p in pdf_files:
            if term in p.name.lower():
                logging.info("Fallback term '%s' matched file '%s'", term, p.name)
                return p

    # 4) final fallback: first pdf
    logging.info("No specific match found. Returning first PDF: %s", pdf_files[0].name)
    return pdf_files[0]


def log_submission_csv(name: str, email_addr: str, mobile: str, pdf_name: Optional[str]) -> None:
    """Append the submission to a CSV file for lead tracking."""
    header = ["timestamp_utc", "name", "email", "mobile", "pdf_requested"]
    first_write = not SUBMISSIONS_CSV.exists()
    row = [datetime.utcnow().isoformat(), name, email_addr, mobile, pdf_name or ""]
    try:
        with SUBMISSIONS_CSV.open("a", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            if first_write:
                writer.writerow(header)
            writer.writerow(row)
    except Exception:
        logging.exception("Failed to write to submissions CSV.")


def log_submission_excel(name: str, email_addr: str, mobile: str, pdf_name: Optional[str]) -> None:
    """
    Append submission to an Excel file (submissions.xlsx) using openpyxl.
    If file doesn't exist, create it with a header row.
    """
    header = ["timestamp_utc", "name", "email", "mobile", "pdf_requested"]
    row = [datetime.utcnow().isoformat(), name, email_addr, mobile, pdf_name or ""]

    try:
        if not SUBMISSIONS_XLSX.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "submissions"
            ws.append(header)
            ws.append(row)
            for idx, col in enumerate(header, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = max(12, len(col) + 2)
            wb.save(SUBMISSIONS_XLSX)
            logging.info("Created submissions.xlsx and logged first row.")
            return

        wb = load_workbook(SUBMISSIONS_XLSX)
        ws = wb.active
        ws.append(row)
        wb.save(SUBMISSIONS_XLSX)
        logging.info("Appended submission to submissions.xlsx")
    except Exception:
        logging.exception("Failed to write to submissions.xlsx")


# ---------- Routes ----------
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "time": datetime.utcnow().isoformat()}), 200


@app.route("/download", methods=["POST"])
def download():
    """
    Accepts JSON:
      { "name": "...", "email": "...", "mobile": "...", "pdf": "optional-key-or-filename" }
    """
    if not request.is_json:
        return jsonify({"message": "Expected JSON body"}), 400
    payload = request.get_json() or {}

    name = safe_str(payload.get("name"))
    email_addr = safe_str(payload.get("email"))
    mobile = safe_str(payload.get("mobile"))
    pdf_key = safe_str(payload.get("pdf"))

    logging.info("Download request: name=%s email=%s mobile=%s pdf_key=%s", name, email_addr, mobile, pdf_key or "(empty)")

    # validate
    err = validate_submission(name, email_addr, mobile)
    if err:
        logging.info("Validation failed: %s", err)
        return jsonify({"message": err}), 400

    # locate PDF
    pdf_path = find_pdf_by_key(pdf_key)
    if not pdf_path or not pdf_path.exists():
        logging.error("PDF not found for key=%s", pdf_key)
        return jsonify({"message": "Requested PDF not found on server."}), 404

    # Log submissions (CSV + Excel)
    try:
        log_submission_csv(name, email_addr, mobile, pdf_path.name)
    except Exception:
        logging.exception("Failed to log to CSV")
    try:
        log_submission_excel(name, email_addr, mobile, pdf_path.name)
    except Exception:
        logging.exception("Failed to log to Excel")

    # respond with the PDF as attachment
    try:
        mimetype, _ = mimetypes.guess_type(str(pdf_path))
        logging.info("Serving file: %s", pdf_path.name)
        return send_file(
            path_or_file=str(pdf_path),
            as_attachment=True,
            download_name=pdf_path.name,
            mimetype=mimetype or "application/pdf"
        )
    except Exception:
        logging.exception("Failed to send file")
        return jsonify({"message": "Internal server error while sending file."}), 500


# ---------- Run ----------
if __name__ == "__main__":
    logging.info("Starting app on %s:%s  DEBUG=%s", HOST, PORT, DEBUG)
    app.run(host=HOST, port=PORT, debug=DEBUG)
